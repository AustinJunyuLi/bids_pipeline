import csv
import re
from collections.abc import Iterable, Mapping
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.config import PROJECT_ROOT, SEEDS_PATH
from pipeline.models.common import PIPELINE_VERSION
from pipeline.models.seed import SeedEvidence, SeedRegistryEntry
from pipeline.models.source import SeedDeal


RAW_SEEDS_XLSX = PROJECT_ROOT / "raw" / "deal_details_Alex_2026.xlsx"
REQUIRED_SEED_COLUMNS = (
    "deal_slug",
    "target_name",
    "acquirer",
    "date_announced",
    "primary_url",
    "is_reference",
)
SEED_COLUMNS = REQUIRED_SEED_COLUMNS

CANONICAL_SLUG_OVERRIDES = {
    "imprivata-inc": "imprivata",
    "medivation-inc": "medivation",
    "zep-inc": "zep",
    "penford-corp": "penford",
    "mac-gray-corp": "mac-gray",
    "saks-inc": "saks",
    "providence-worcester-rr-co": "providence-worcester",
    "stec-inc": "stec",
    "s-tec-inc": "stec",
    "s-t-e-c-inc": "stec",
}
REFERENCE_SLUGS = {
    "imprivata",
    "medivation",
    "zep",
    "petsmart-inc",
    "penford",
    "mac-gray",
    "saks",
    "providence-worcester",
    "stec",
}
XLSX_COLUMN_MAP = {
    "TargetName": "target_name",
    "Acquirer": "acquirer",
    "DateAnnounced": "date_announced",
    "URL": "primary_url",
}


def make_slug(name: str) -> str:
    """Lowercase, strip punctuation, collapse to hyphens."""

    slug = re.sub(r"[^a-z0-9]+", "-", name.lower())
    slug = re.sub(r"-{2,}", "-", slug)
    return slug.strip("-")


def canonical_deal_slug(target_name: str) -> str:
    """Apply stable overrides for known reference/example deals."""

    base_slug = make_slug(target_name)
    return CANONICAL_SLUG_OVERRIDES.get(base_slug, base_slug)


def validate_seed_headers(headers: Iterable[str]) -> None:
    header_set = {header for header in headers if header}
    missing = [column for column in REQUIRED_SEED_COLUMNS if column not in header_set]
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Seed input is missing required columns: {missing_text}")


def load_seed_registry(path: Path = SEEDS_PATH) -> list[SeedRegistryEntry]:
    rows = _load_rows(path)
    return build_seed_registry(rows)


def build_seed_registry(rows: Iterable[Mapping[str, Any]]) -> list[SeedRegistryEntry]:
    entries_by_slug: dict[str, SeedRegistryEntry] = {}
    ordered_slugs: list[str] = []

    for row in rows:
        target_name = _clean_optional_text(row.get("target_name"))
        if not target_name:
            continue

        deal_slug = canonical_deal_slug(target_name)
        entry = entries_by_slug.get(deal_slug)
        is_reference = _parse_bool(row.get("is_reference")) or deal_slug in REFERENCE_SLUGS
        acquirer = _clean_optional_text(row.get("acquirer"))
        date_announced = _parse_seed_date(row.get("date_announced"))
        primary_url = _clean_optional_text(row.get("primary_url"))

        if entry is None:
            entry = SeedRegistryEntry(
                deal_slug=deal_slug,
                target_name=target_name,
                acquirer=acquirer,
                date_announced=date_announced,
                primary_url=primary_url,
                is_reference=is_reference,
            )
            entries_by_slug[deal_slug] = entry
            ordered_slugs.append(deal_slug)
        else:
            entry.is_reference = entry.is_reference or is_reference
            if entry.acquirer is None and acquirer is not None:
                entry.acquirer = acquirer
            if entry.date_announced is None and date_announced is not None:
                entry.date_announced = date_announced
            if entry.primary_url is None and primary_url is not None:
                entry.primary_url = primary_url

        entry.evidence.append(
            SeedEvidence(
                row_ref=_clean_optional_text(row.get("_row_ref")),
                provided_deal_slug=_clean_optional_text(row.get("deal_slug")),
                target_name_raw=target_name,
                acquirer_raw=acquirer,
                date_announced_raw=_clean_optional_text(row.get("date_announced")),
                primary_url_raw=primary_url,
                is_reference_raw=_parse_bool(row.get("is_reference")),
            )
        )
        _record_conflict(entry, "target_name", target_name)
        _record_conflict(entry, "acquirer", acquirer)
        _record_conflict(entry, "date_announced", _clean_optional_text(row.get("date_announced")))
        _record_conflict(entry, "primary_url", primary_url)

    return [entries_by_slug[slug] for slug in ordered_slugs]


def entry_to_seed_artifact(
    entry: SeedRegistryEntry,
    *,
    run_id: str,
    pipeline_version: str = PIPELINE_VERSION,
) -> SeedDeal:
    row_refs = [evidence.row_ref for evidence in entry.evidence if evidence.row_ref]
    notes = [
        f"conflict:{field}=" + " | ".join(values)
        for field, values in sorted(entry.conflicting_evidence.items())
    ]
    return SeedDeal(
        run_id=run_id,
        pipeline_version=pipeline_version,
        deal_slug=entry.deal_slug,
        target_name=entry.target_name,
        acquirer_seed=entry.acquirer,
        date_announced_seed=entry.date_announced,
        primary_url_seed=entry.primary_url,
        is_reference=entry.is_reference,
        seed_row_refs=row_refs,
        notes=notes,
    )


def extract_seeds(xlsx_path: Path) -> list[dict[str, str]]:
    """Read deal_details xlsx and return normalized seed rows."""

    return [_entry_to_output_row(entry) for entry in load_seed_registry(xlsx_path)]


def write_seeds(
    seeds: Iterable[SeedRegistryEntry | SeedDeal | Mapping[str, Any]],
    output_path: Path,
) -> None:
    """Write seeds to CSV."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=SEED_COLUMNS)
        writer.writeheader()
        for seed in seeds:
            writer.writerow(_coerce_output_row(seed))


def main() -> None:
    """Entry point: rebuild seeds.csv from workbook when available, else normalize the CSV."""

    input_path = RAW_SEEDS_XLSX if RAW_SEEDS_XLSX.exists() else SEEDS_PATH
    entries = load_seed_registry(input_path)
    write_seeds(entries, SEEDS_PATH)
    print(f"Wrote {len(entries)} seeds to {SEEDS_PATH}")


def _coerce_output_row(seed: SeedRegistryEntry | SeedDeal | Mapping[str, Any]) -> dict[str, str]:
    if isinstance(seed, SeedRegistryEntry):
        return _entry_to_output_row(seed)
    if isinstance(seed, SeedDeal):
        return {
            "deal_slug": seed.deal_slug,
            "target_name": seed.target_name,
            "acquirer": seed.acquirer_seed or "",
            "date_announced": seed.date_announced_seed.isoformat()
            if seed.date_announced_seed
            else "",
            "primary_url": seed.primary_url_seed or "",
            "is_reference": "true" if seed.is_reference else "false",
        }

    return {
        "deal_slug": _clean_optional_text(seed.get("deal_slug")) or "",
        "target_name": _clean_optional_text(seed.get("target_name")) or "",
        "acquirer": _clean_optional_text(seed.get("acquirer")) or "",
        "date_announced": _clean_optional_text(seed.get("date_announced")) or "",
        "primary_url": _clean_optional_text(seed.get("primary_url")) or "",
        "is_reference": "true" if _parse_bool(seed.get("is_reference")) else "false",
    }


def _entry_to_output_row(entry: SeedRegistryEntry) -> dict[str, str]:
    return {
        "deal_slug": entry.deal_slug,
        "target_name": entry.target_name,
        "acquirer": entry.acquirer or "",
        "date_announced": entry.date_announced.isoformat() if entry.date_announced else "",
        "primary_url": entry.primary_url or "",
        "is_reference": "true" if entry.is_reference else "false",
    }


def _load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(path)
    raise ValueError(f"Unsupported seed input format: {path.suffix}")


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        validate_seed_headers(fieldnames)
        rows: list[dict[str, Any]] = []
        for line_no, row in enumerate(reader, start=2):
            enriched = {key: value for key, value in row.items()}
            enriched["_row_ref"] = f"csv:{line_no}"
            rows.append(enriched)
        return rows


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_rows = worksheet.iter_rows(values_only=True)
    header = [str(value).strip() if value else "" for value in next(raw_rows)]
    indices = {name: idx for idx, name in enumerate(header) if name}
    missing = [name for name in XLSX_COLUMN_MAP if name not in indices]
    if missing:
        raise ValueError(
            "Seed workbook is missing required columns: " + ", ".join(sorted(missing))
        )

    rows: list[dict[str, Any]] = []
    for row_no, values in enumerate(raw_rows, start=2):
        target_name = _clean_optional_text(values[indices["TargetName"]])
        if not target_name:
            continue
        rows.append(
            {
                "deal_slug": canonical_deal_slug(target_name),
                "target_name": target_name,
                "acquirer": _clean_optional_text(values[indices["Acquirer"]]),
                "date_announced": _format_date(values[indices["DateAnnounced"]]),
                "primary_url": _clean_optional_text(values[indices["URL"]]),
                "is_reference": canonical_deal_slug(target_name) in REFERENCE_SLUGS,
                "_row_ref": f"xlsx:{row_no}",
            }
        )
    return rows


def _record_conflict(entry: SeedRegistryEntry, field_name: str, candidate: str | None) -> None:
    if not candidate:
        return

    existing_values = set(entry.conflicting_evidence.get(field_name, []))
    canonical_value = _canonical_field_text(entry, field_name)
    if canonical_value:
        existing_values.add(canonical_value)
    existing_values.add(candidate)

    if len(existing_values) > 1:
        entry.conflicting_evidence[field_name] = sorted(existing_values)


def _canonical_field_text(entry: SeedRegistryEntry, field_name: str) -> str | None:
    value = getattr(entry, field_name)
    if isinstance(value, date):
        return value.isoformat()
    return value


def _clean_optional_text(value: object) -> str | None:
    cleaned = _clean_text(value)
    return cleaned or None


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = _clean_text(value).lower()
    return text in {"1", "true", "t", "yes", "y"}


def _parse_seed_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _clean_text(value)
    if not text:
        return None

    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _format_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _clean_text(value)


if __name__ == "__main__":
    main()
