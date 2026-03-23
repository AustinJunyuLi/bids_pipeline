from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]

DROP_LABEL_NORMALIZATION = {
    "DropBelowM": "DropM",
    "DropM": "DropM",
    "Drop": "Drop",
    "DropBelowInf": "DropBelowInf",
    "DropAtInf": "DropAtInf",
    "DropTarget": "DropTarget",
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python scripts/reconcile_alex.py",
        description="Standalone post-production reconciliation against Alex's workbook.",
    )
    parser.add_argument("--deal", required=True)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    return parser


def _normalize_company_name(name: str) -> str:
    normalized = re.sub(r"[^a-z0-9 ]", " ", name.strip().lower())
    for suffix in ("inc", "corp", "corporation", "llc", "ltd", "l p", "co", "company"):
        normalized = re.sub(rf"\b{suffix}\b", "", normalized)
    return " ".join(normalized.split())


def _normalize_actor_name(name: str | None) -> str | None:
    if not name:
        return None
    normalized = re.sub(r"[^a-z0-9 ]", " ", name.strip().lower())
    normalized = re.sub(r"^\bthe\b\s+", "", normalized)
    for suffix in (
        "inc",
        "corp",
        "corporation",
        "llc",
        "ltd",
        "l p",
        "co",
        "company",
    ):
        normalized = re.sub(rf"\b{suffix}\b", "", normalized)
    return " ".join(normalized.split())


def _require(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required input: {path}")


def _load_seed_target_name(project_root: Path, slug: str) -> str:
    seeds_path = project_root / "data" / "seeds.csv"
    _require(seeds_path)
    with seeds_path.open(encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            if row["deal_slug"] == slug:
                return row["target_name"]
    raise FileNotFoundError(f"Deal slug {slug!r} not found in {seeds_path}")


def _extract_alex_rows(project_root: Path, slug: str, target_name: str) -> list[dict[str, Any]]:
    workbook_path = project_root / "example" / "deal_details_Alex_2026.xlsx"
    _require(workbook_path)

    normalized_target = _normalize_company_name(target_name)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["deal_details"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, values, strict=False))
            raw_target = record.get("TargetName")
            if not raw_target or _normalize_company_name(str(raw_target)) != normalized_target:
                continue
            serializable: dict[str, Any] = {}
            for key, value in record.items():
                if isinstance(value, datetime):
                    serializable[key] = value.isoformat()
                else:
                    serializable[key] = value
            rows.append(serializable)
    finally:
        workbook.close()

    if not rows:
        raise FileNotFoundError(
            f"No Alex rows found in {workbook_path} for target {target_name!r}"
        )

    reconcile_dir = project_root / "data" / "skill" / slug / "reconcile"
    reconcile_dir.mkdir(parents=True, exist_ok=True)
    (reconcile_dir / "alex_rows.json").write_text(
        json.dumps(rows, indent=2, default=str),
        encoding="utf-8",
    )
    return rows


def _load_export_rows(project_root: Path, slug: str) -> list[dict[str, str]]:
    export_path = project_root / "data" / "skill" / slug / "export" / "deal_events.csv"
    _require(export_path)
    lines = export_path.read_text(encoding="utf-8").splitlines()
    if len(lines) < 4:
        return []
    data_lines = lines[3:]
    if not data_lines:
        return []
    reader = csv.DictReader(data_lines)
    return list(reader)


def _parse_export_date(value: str | None) -> str | None:
    if not value or value == "NA":
        return None
    return datetime.strptime(value, "%m/%d/%Y").date().isoformat()


def _parse_alex_date(value: Any) -> str | None:
    if value in (None, "", "NA"):
        return None
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).date().isoformat()
        except ValueError:
            return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return None


def _is_populated_amount(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip().upper() in {"", "NA"}:
        return False
    return True


def _pipeline_record(row: dict[str, str], index: int) -> dict[str, Any]:
    note = row.get("note", "")
    value = row.get("val", "")
    family = note
    subtype = None
    if note == "NA" and value not in {"", "NA"}:
        family = "proposal"
    elif note in DROP_LABEL_NORMALIZATION:
        family = "drop"
        subtype = DROP_LABEL_NORMALIZATION[note]
    return {
        "record_id": f"pipeline:{index}",
        "family": family,
        "subtype": subtype,
        "actor": _normalize_actor_name(row.get("bidder")),
        "date": _parse_export_date(row.get("date_p") or row.get("date_r")),
        "value": row.get("val"),
        "bid_type": row.get("bid_type"),
        "raw": row,
    }


def _alex_record(row: dict[str, Any], index: int) -> dict[str, Any]:
    bid_note = row.get("bid_note")
    family = bid_note
    subtype = None
    if bid_note == "NA" and any(
        _is_populated_amount(row.get(key))
        for key in ("bid_value", "bid_value_pershare", "bid_value_lower", "bid_value_upper")
    ):
        family = "proposal"
    elif bid_note in DROP_LABEL_NORMALIZATION:
        family = "drop"
        subtype = DROP_LABEL_NORMALIZATION[bid_note]
    return {
        "record_id": f"alex:{index}",
        "family": family,
        "subtype": subtype,
        "actor": _normalize_actor_name(row.get("BidderName")),
        "date": _parse_alex_date(row.get("bid_date_rough")),
        "value": row.get("bid_value_pershare") or row.get("bid_value"),
        "bid_type": row.get("bid_type"),
        "raw": row,
    }


def _compare_matched_pair(pipeline: dict[str, Any], alex: dict[str, Any]) -> list[dict[str, Any]]:
    mismatches: list[dict[str, Any]] = []
    if pipeline["family"] == "proposal":
        pipeline_value = pipeline["raw"].get("val")
        alex_value = alex["value"]
        if pipeline_value not in {None, "", "NA"} and alex_value not in {None, "", "NA"}:
            if abs(float(pipeline_value) - float(alex_value)) > 0.01:
                mismatches.append(
                    {
                        "pipeline_event_id": pipeline["record_id"],
                        "alex_bidder_id": alex["raw"].get("BidderID"),
                        "field": "bid_value",
                        "pipeline_value": str(pipeline_value),
                        "alex_value": str(alex_value),
                        "filing_supports": "no_evidence",
                        "filing_snippet": None,
                    }
                )
        pipeline_bid_type = (pipeline["bid_type"] or "").strip()
        alex_bid_type = (alex["bid_type"] or "").strip()
        if pipeline_bid_type and alex_bid_type and pipeline_bid_type != alex_bid_type:
            mismatches.append(
                {
                    "pipeline_event_id": pipeline["record_id"],
                    "alex_bidder_id": alex["raw"].get("BidderID"),
                    "field": "bid_type",
                    "pipeline_value": pipeline_bid_type,
                    "alex_value": alex_bid_type,
                    "filing_supports": "no_evidence",
                    "filing_snippet": None,
                }
            )
    if pipeline["family"] == "drop":
        if pipeline["subtype"] and alex["subtype"] and pipeline["subtype"] != alex["subtype"]:
            mismatches.append(
                {
                    "pipeline_event_id": pipeline["record_id"],
                    "alex_bidder_id": alex["raw"].get("BidderID"),
                    "field": "dropout_subtype",
                    "pipeline_value": pipeline["subtype"],
                    "alex_value": alex["subtype"],
                    "filing_supports": "no_evidence",
                    "filing_snippet": None,
                }
            )
    return mismatches


def _build_orphan(source: str, record: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": source,
        "event_id": record["record_id"] if source == "pipeline" else None,
        "alex_bidder_id": record["raw"].get("BidderID") if source == "alex" else None,
        "event_type": record["family"],
        "actor_name": record["raw"].get("bidder") if source == "pipeline" else record["raw"].get("BidderName"),
        "date_rough": record["date"],
        "value": record["value"],
        "filing_verdict": "ungrounded",
        "filing_snippet": None,
    }


def _match_records(
    pipeline_records: list[dict[str, Any]],
    alex_records: list[dict[str, Any]],
) -> tuple[list[tuple[dict[str, Any], dict[str, Any]]], list[dict[str, Any]], list[dict[str, Any]]]:
    matches: list[tuple[dict[str, Any], dict[str, Any]]] = []
    unmatched_pipeline = list(pipeline_records)
    unmatched_alex = list(alex_records)

    pipeline_by_key: dict[tuple[Any, Any, Any], dict[str, Any]] = {}
    for record in pipeline_records:
        pipeline_by_key[(record["family"], record["date"], record["actor"])] = record

    used_pipeline_ids: set[str] = set()
    used_alex_ids: set[str] = set()
    for alex in alex_records:
        key = (alex["family"], alex["date"], alex["actor"])
        candidate = pipeline_by_key.get(key)
        if candidate is None or candidate["record_id"] in used_pipeline_ids:
            continue
        matches.append((candidate, alex))
        used_pipeline_ids.add(candidate["record_id"])
        used_alex_ids.add(alex["record_id"])

    unmatched_pipeline = [
        record for record in pipeline_records if record["record_id"] not in used_pipeline_ids
    ]
    unmatched_alex = [
        record for record in alex_records if record["record_id"] not in used_alex_ids
    ]
    return matches, unmatched_pipeline, unmatched_alex


def _summary_status(
    *,
    matched_count: int,
    pipeline_event_count: int,
    alex_event_count: int,
    field_mismatch_count: int,
    orphan_ungrounded_count: int,
) -> str:
    denominator = max(pipeline_event_count, alex_event_count)
    match_rate = 1.0 if denominator == 0 else matched_count / denominator
    if match_rate < 0.7 or orphan_ungrounded_count > 0:
        return "fail" if match_rate < 0.7 else "warn"
    if match_rate < 0.9 or field_mismatch_count > 0:
        return "warn"
    return "pass"


def run_reconcile(deal_slug: str, *, project_root: Path = PROJECT_ROOT) -> int:
    skill_root = project_root / "data" / "skill" / deal_slug
    reconcile_dir = skill_root / "reconcile"
    reconcile_dir.mkdir(parents=True, exist_ok=True)

    required_paths = [
        skill_root / "extract" / "actors_raw.json",
        skill_root / "extract" / "events_raw.json",
        skill_root / "extract" / "spans.json",
        skill_root / "export" / "deal_events.csv",
        project_root / "data" / "deals" / deal_slug / "source" / "chronology_blocks.jsonl",
        project_root / "raw" / deal_slug / "document_registry.json",
    ]
    for path in required_paths:
        _require(path)
    enrich_path = skill_root / "enrich" / "enrichment.json"
    deterministic_path = skill_root / "enrich" / "deterministic_enrichment.json"
    if not enrich_path.exists() and not deterministic_path.exists():
        raise FileNotFoundError(
            f"Missing required input: {enrich_path} or {deterministic_path}"
        )

    target_name = _load_seed_target_name(project_root, deal_slug)
    alex_rows = _extract_alex_rows(project_root, deal_slug, target_name)
    export_rows = _load_export_rows(project_root, deal_slug)

    pipeline_records = [
        _pipeline_record(row, index)
        for index, row in enumerate(export_rows, start=1)
    ]
    alex_records = [
        _alex_record(row, index)
        for index, row in enumerate(alex_rows, start=1)
    ]

    matches, pipeline_only, alex_only = _match_records(pipeline_records, alex_records)
    field_mismatches: list[dict[str, Any]] = []
    for pipeline_record, alex_record in matches:
        field_mismatches.extend(_compare_matched_pair(pipeline_record, alex_record))

    pipeline_only_payload = [_build_orphan("pipeline", record) for record in pipeline_only]
    alex_only_payload = [_build_orphan("alex", record) for record in alex_only]

    denominator = max(len(pipeline_records), len(alex_records))
    match_rate = 1.0 if denominator == 0 else len(matches) / denominator
    orphan_ungrounded_count = len(pipeline_only_payload) + len(alex_only_payload)
    summary = {
        "status": _summary_status(
            matched_count=len(matches),
            pipeline_event_count=len(pipeline_records),
            alex_event_count=len(alex_records),
            field_mismatch_count=len(field_mismatches),
            orphan_ungrounded_count=orphan_ungrounded_count,
        ),
        "match_rate": match_rate,
        "field_mismatch_count": len(field_mismatches),
        "arbitration_pipeline_wins": 0,
        "arbitration_alex_wins": 0,
        "arbitration_inconclusive": len(field_mismatches),
        "orphan_grounded_count": 0,
        "orphan_ungrounded_count": orphan_ungrounded_count,
    }
    report = {
        "deal_slug": deal_slug,
        "target_name": target_name,
        "pipeline_event_count": len(pipeline_records),
        "alex_event_count": len(alex_records),
        "matched_count": len(matches),
        "pipeline_only_count": len(pipeline_only_payload),
        "alex_only_count": len(alex_only_payload),
        "aggregate_mismatch_count": 0,
        "field_mismatches": field_mismatches,
        "pipeline_only": pipeline_only_payload,
        "alex_only": alex_only_payload,
        "summary": summary,
    }

    (reconcile_dir / "reconciliation_report.json").write_text(
        json.dumps(report, indent=2),
        encoding="utf-8",
    )
    return 1 if summary["status"] == "fail" else 0


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    return run_reconcile(args.deal, project_root=args.project_root)


if __name__ == "__main__":
    raise SystemExit(main())
